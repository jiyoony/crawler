#-*- coding: utf-8 -*-
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

list_name = []
for i in range(1,1000):
    #페이지 넘기면서 반복
    req =requests.get('https://www.instiz.net/pt?page='+str(i))
    soup =BeautifulSoup(req.text,'html.parser')

# print(soup.find_all("td",class_="listnm"))

    for i in soup.find_all("td",class_="minitext2 listnm"):
        list_name.append(i.text)
        # print(i.text)

count = {}
# 딕셔너리 사용
for i in list_name:
    try: count[i] += 1
    except: count[i]=1

id = list(count.keys())
cnt = list(count.values())
print(id)
print(cnt)
write_wb =Workbook()
write_ws =write_wb.active
#
for i in range(1,len(id)):
    # write_ws.cell(i,1,list_name[i-1])
    write_ws.cell(i,1,id[i-1])
    write_ws.cell(i,2,cnt[i-1])

write_wb.save("writer20.xlsx")
